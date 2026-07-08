import io
from datetime import datetime
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
BLUE_DARK    = "0C447C"
BLUE_MID     = "185FA5"
BLUE_LIGHT   = "E6F1FB"
GREEN_DARK   = "27500A"
GREEN_MID    = "3B6D11"
GREEN_LIGHT  = "EAF3DE"
PURPLE_DARK  = "3C3489"
PURPLE_MID   = "534AB7"
PURPLE_LIGHT = "EEEDFE"
WHITE        = "FFFFFF"
GRAY_LIGHT   = "F1EFE8"
GRAY_DARK    = "2C2C2A"
AMBER        = "FAEEDA"
AMBER_DARK   = "854F0B"

COPYRIGHT         = "© 2026 News Digest Platform. Internal use only. All rights reserved."
ARTICLE_ROW_HEIGHT = 65


# ── Style helpers ──────────────────────────────────────────────────────────────
def _fill(hex_color: str | None) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color or GRAY_DARK)

def _font(bold=False, size=10, color=GRAY_DARK, underline=None) -> Font:
    return Font(name="Calibri", bold=bold, size=size,
                color=color, underline=underline)

def _border() -> Border:
    s = Side(style="thin", color="D3D1C7")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="top", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Shared row writers ─────────────────────────────────────────────────────────
def _title_row(ws, text: str, ncols: int, bg: str, row=1):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(bold=True, size=14, color=WHITE)
    c.fill      = _fill(bg) if bg else _fill(GRAY_DARK)
    c.alignment = _center()
    ws.row_dimensions[row].height = 28

def _period_row(ws, window: dict | None, ncols: int, bg: str = "", row: int = 2):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    label = (window or {}).get("label", "")
    c = ws.cell(row=row, column=1,
                value=f"Period: {label}")
    c.font      = _font(bold=False, size=11, color=WHITE)
    c.fill      = _fill(bg)
    c.alignment = _center()
    ws.row_dimensions[row].height = 18

def _header_row(ws, headers: list, row: int, bg: str):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = _font(bold=True, size=10, color=WHITE)
        c.fill      = _fill(bg)
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[row].height = 22

def _data_cell(ws, row: int, col: int, value, fill: PatternFill,
               hyperlink: str | None = None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill
    c.border    = _border()
    if hyperlink:
        c.hyperlink = hyperlink
        c.font      = _font(size=10, color=BLUE_MID, underline="single")
        c.alignment = _align()
    else:
        c.font      = _font(size=10)
        c.alignment = _align()
    return c

def _set_widths(ws, widths: list):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def _copyright_row(ws, ncols: int, row: int):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=COPYRIGHT)
    c.font      = _font(size=9, color="888880")
    c.alignment = _center()
    ws.row_dimensions[row].height = 14

def _parse_date(date_str: str):
    try:
        dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
        return dt.strftime("%B"), dt.strftime("%B %Y"), dt.strftime("%d %b %Y")
    except Exception:
        return "", "", date_str


# ── Summary sheet ──────────────────────────────────────────────────────────────
def _build_summary(wb, client_data, prospect_data,
                   industry_data, window: dict | None):
    ws = wb.create_sheet("Summary", 0)
    _set_widths(ws, [32, 20])

    _title_row(ws, "News Digest Platform — Export Summary", 2, GRAY_DARK, row=1)

    ws.merge_cells("A2:B2")
    c = ws.cell(row=2, column=1,
                value=f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}")
    c.font = _font(size=10, color="5F5E5A")
    c.alignment = _center()

    ws.merge_cells("A3:B3")
    label = window.get("label", "") if window else ""
    c = ws.cell(row=3, column=1, value=f"Period: {label}")
    c.font = _font(bold=True, size=10, color=GRAY_DARK)
    c.alignment = _center()
    ws.row_dimensions[3].height = 16

    rows = [
        ("Category",    "Articles",  True),
        ("Clients",
         sum(len(i["news"]) for i in client_data),   False),
        ("Prospects",
         sum(len(i["news"]) for i in prospect_data), False),
        ("Industries",
         sum(len(i["news"]) for i in industry_data), False),
        ("",            "",          False),
        ("Total articles exported",
         sum(len(i["news"])
             for i in client_data + prospect_data + industry_data), True),
    ]
    for ri, (label, value, bold) in enumerate(rows, 4):
        ca = ws.cell(row=ri, column=1, value=label)
        cb = ws.cell(row=ri, column=2, value=value)
        fill = _fill(GRAY_LIGHT) if bold else _fill(WHITE)
        for c in (ca, cb):
            c.font      = _font(bold=bold, size=10)
            c.fill      = fill
            c.alignment = _align("left" if c == ca else "center")
            if label:
                c.border = _border()
        ws.row_dimensions[ri].height = 18

    _copyright_row(ws, 2, len(rows) + 5)
    ws.freeze_panes = "A4"


# ── Sheet setup helper ─────────────────────────────────────────────────────────
def _sheet_header(ws, title: str, headers: list, window, dark_color: str, mid_color: str) -> int:
    """Write title / period / column-header rows and return the column count."""
    ncols = len(headers)
    _title_row(ws,  title,   ncols, dark_color)
    _period_row(ws, window,  ncols, mid_color)
    _header_row(ws, headers, row=3, bg=mid_color)
    return ncols


# ── Client Report sheet ────────────────────────────────────────────────────────
def _build_client_sheet(wb, client_data: List[Dict], window: dict | None):
    """
    Columns: Company Name | News Article Header | News Article Details
             (with Insights) | Source | Date | Month | Primary Category |
             Fetch Date | Period | Fetch API | Hyperlink | Client/Prospect
    """
    ws      = wb.create_sheet("Client Report")
    headers = [
        "Company Name", "News Article Header",
        "News Article Details (with Insights)",
        "Source", "Date", "Month", "Primary Category",
        "Fetch Date", "Period", "Fetch API",
        "Hyperlink", "Client/Prospect",
    ]
    ncols = _sheet_header(ws, "Client News Report", headers, window, BLUE_DARK, BLUE_MID)
    _set_widths(ws, [24, 36, 55, 20, 14, 14, 26, 14, 26, 12, 14, 16])

    row = 4
    for item in client_data:
        entity = item["entity"]
        for idx, n in enumerate(item["news"]):
            month, _, date_fmt = _parse_date(n.published_date)
            fill = _fill(BLUE_LIGHT) if idx % 2 == 0 else _fill(WHITE)

            _data_cell(ws, row,  1, entity.name,                     fill)
            _data_cell(ws, row,  2, n.title,                         fill)
            _data_cell(ws, row,  3, n.summary or n.title,            fill)
            _data_cell(ws, row,  4, n.source,                        fill)
            _data_cell(ws, row,  5, date_fmt,                        fill)
            _data_cell(ws, row,  6, month,                           fill)
            _data_cell(ws, row,  7, n.primary_category or "",        fill)
            _data_cell(ws, row,  8, n.fetched_date or "",            fill)
            _data_cell(ws, row,  9, n.period or "",                  fill)
            _data_cell(ws, row, 10, n.fetch_source or "",            fill)
            link_label = "⚠ View Article" if n.paywall_note else "View Article"
            _data_cell(ws, row, 11, link_label,                      fill,
                       hyperlink=n.url or n.original_url)
            _data_cell(ws, row, 12, "Client",                        fill)

            ws.row_dimensions[row].height = ARTICLE_ROW_HEIGHT
            row += 1

    _copyright_row(ws, ncols, row + 1)
    ws.freeze_panes = "A4"


# ── Prospect Report sheet ──────────────────────────────────────────────────────
def _build_prospect_sheet(wb, prospect_data: List[Dict], window: dict | None):
    """
    Columns: Company | News Article Header | News Article Details |
             Source | Month, Year | Date | Primary Category |
             Fetch Date | Period | Fetch API | News HyperLink | Client/Prospect
    """
    ws      = wb.create_sheet("Prospect Report")
    headers = [
        "Company", "News Article Header", "News Article Details",
        "Source", "Month, Year", "Date", "Primary Category",
        "Fetch Date", "Period", "Fetch API",
        "News HyperLink", "Client/Prospect",
    ]
    ncols = _sheet_header(ws, "Prospect News Report", headers, window, GREEN_DARK, GREEN_MID)
    _set_widths(ws, [24, 36, 55, 20, 16, 14, 26, 14, 26, 12, 14, 16])

    row = 4
    for item in prospect_data:
        entity = item["entity"]
        for idx, n in enumerate(item["news"]):
            month, month_year, date_fmt = _parse_date(n.published_date)
            fill = _fill(GREEN_LIGHT) if idx % 2 == 0 else _fill(WHITE)

            _data_cell(ws, row,  1, entity.name,                  fill)
            _data_cell(ws, row,  2, n.title,                      fill)
            _data_cell(ws, row,  3, n.summary or n.title,         fill)
            _data_cell(ws, row,  4, n.source,                     fill)
            _data_cell(ws, row,  5, month_year,                   fill)
            _data_cell(ws, row,  6, date_fmt,                     fill)
            _data_cell(ws, row,  7, n.primary_category or "",     fill)
            _data_cell(ws, row,  8, n.fetched_date or "",         fill)
            _data_cell(ws, row,  9, n.period or "",               fill)
            _data_cell(ws, row, 10, n.fetch_source or "",         fill)
            _data_cell(ws, row, 11, "View Article",               fill,
                       hyperlink=n.url or n.original_url)
            _data_cell(ws, row, 12, "Prospect",                   fill)

            ws.row_dimensions[row].height = ARTICLE_ROW_HEIGHT
            row += 1

    _copyright_row(ws, ncols, row + 1)
    ws.freeze_panes = "A4"


# ── Industry News sheet ────────────────────────────────────────────────────────
def _build_industry_sheet(wb, industry_data: List[Dict], window: dict | None):
    """
    Columns: Industry | Primary Category | Secondary Category |
             News Article Header | News Article Details (with Insights) |
             Source | Article Date | Month | Fetch Date | Period | Fetch API | Hyperlink
    """
    ws      = wb.create_sheet("Industry News")
    headers = [
        "Industry",
        "Primary Category", "Secondary Category",
        "News Article Header",
        "News Article Details (with Insights)",
        "Source", "Article Date", "Month",
        "Fetch Date", "Period", "Fetch API",
        "Hyperlink",
    ]
    ncols = _sheet_header(ws, "Industry News Report", headers, window, PURPLE_DARK, PURPLE_MID)
    _set_widths(ws, [22, 30, 28, 36, 55, 20, 14, 14, 14, 26, 12, 14])

    row = 4
    for item in industry_data:
        entity = item["entity"]
        for idx, n in enumerate(item["news"]):
            month, _, date_fmt = _parse_date(n.published_date)
            fill = _fill(PURPLE_LIGHT) if idx % 2 == 0 else _fill(WHITE)

            _data_cell(ws, row,  1, entity.name,                        fill)
            _data_cell(ws, row,  2, n.primary_category or "General",    fill)
            _data_cell(ws, row,  3, n.secondary_category or "",         fill)
            _data_cell(ws, row,  4, n.title,                            fill)
            _data_cell(ws, row,  5, n.summary or n.title,               fill)
            _data_cell(ws, row,  6, n.source,                           fill)
            _data_cell(ws, row,  7, date_fmt,                           fill)
            _data_cell(ws, row,  8, month,                              fill)
            _data_cell(ws, row,  9, n.fetched_date or "",               fill)
            _data_cell(ws, row, 10, n.period or "",                     fill)
            _data_cell(ws, row, 11, n.fetch_source or "",               fill)
            _data_cell(ws, row, 12, "View Article",                     fill,
                       hyperlink=n.url or n.original_url)

            ws.row_dimensions[row].height = ARTICLE_ROW_HEIGHT
            row += 1

    _copyright_row(ws, ncols, row + 1)
    ws.freeze_panes = "A4"


# ── Gap Report sheet ───────────────────────────────────────────────────────────
def _build_gap_sheet(wb, gap_report: dict):
    """
    Lists all entities with no news at all, and per-entity topic gaps.
    The gap_report dict already contains the window from build_gap_report().
    """
    if not gap_report:
        return

    ws    = wb.create_sheet("Gap Report")
    ncols = 3
    _title_row(ws, "News Gap Report", ncols, AMBER_DARK)
    _period_row(ws, gap_report.get("window"), ncols, AMBER_DARK, row=2)
    _set_widths(ws, [28, 20, 45])

    # Section 1 — no news at all (shifted down one row for period row)
    ws.merge_cells(f"A4:C4")
    c = ws.cell(row=4, column=1, value="Entities with NO news in this period")
    c.font = _font(bold=True, size=10, color=AMBER_DARK)
    c.fill = _fill(AMBER)
    c.alignment = _align()

    _header_row(ws, ["Entity Name", "Type", "Reason"], row=5, bg=AMBER_DARK)

    row = 6
    for item in gap_report.get("no_news_at_all", []):
        fill = _fill(AMBER) if row % 2 == 0 else _fill(WHITE)
        _data_cell(ws, row, 1, item["name"], fill)
        _data_cell(ws, row, 2, item["type"], fill)
        _data_cell(ws, row, 3, "No articles found across all 12 topics", fill)
        ws.row_dimensions[row].height = 18
        row += 1

    # Section 2 — topic gaps
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1, value="Entities with topic-level gaps")
    c.font = _font(bold=True, size=10, color=AMBER_DARK)
    c.fill = _fill(AMBER)
    c.alignment = _align()
    row += 1

    _header_row(ws, ["Entity Name", "Type", "Topics with no articles"], row=row, bg=AMBER_DARK)
    row += 1

    for name, info in gap_report.get("topic_gaps", {}).items():
        fill = _fill(AMBER) if row % 2 == 0 else _fill(WHITE)
        missing = ", ".join(info["missing_topics"])
        _data_cell(ws, row, 1, name,               fill)
        _data_cell(ws, row, 2, info["entity_type"], fill)
        _data_cell(ws, row, 3, missing,             fill)
        ws.row_dimensions[row].height = 40
        row += 1

    _copyright_row(ws, ncols, row + 1)
    ws.freeze_panes = "A6"


# ── Full analytics sheet ───────────────────────────────────────────────────────
def _build_full_analytics(wb, entities_map: dict, all_news: dict,
                          audit_entries: list, window: dict | None = None):
    """
    Single sheet with every article (accepted + rejected) including reason.
    Columns: Entity | Type | Status | Article Title | Source | URL |
             Published | Period | Primary Category | Secondary Category |
             Topic | Fetch Source | Reason / Note
    """
    RED_LIGHT    = "FEE2E2"
    RED_MID      = "B91C1C"
    ORANGE_LIGHT = "FEF9C3"
    ORANGE_MID   = "B45309"
    GREEN_LIGHT2 = "DCFCE7"
    GREEN_MID2   = "166534"

    ws      = wb.create_sheet("Full Analytics")
    headers = [
        "Entity Name", "Entity Type", "Status", "Reason / Note",
        "Article Title", "Source", "URL",
        "Published Date", "Period",
        "Primary Category", "Secondary Category",
        "Topic Queried", "Fetch Source",
    ]
    ncols        = len(headers)
    period_label = (window or {}).get("label", "All time")
    _title_row(ws, "Full Analytics — All Articles (Accepted + Rejected)", ncols, GRAY_DARK)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1,
                value=f"Period: {period_label}  |  "
                      f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  "
                      f"Accepted: {sum(len(v) for v in all_news.values())}  |  "
                      f"Audit entries: {len(audit_entries)}")
    c.font      = _font(size=10, color=WHITE)
    c.fill      = _fill(GRAY_DARK)
    c.alignment = _center()

    _header_row(ws, headers, row=3, bg=GRAY_DARK)
    _set_widths(ws, [22, 12, 14, 40, 44, 20, 14, 14, 28, 30, 28, 28, 14])

    row = 4

    # ── Accepted articles from news data ──────────────────────────────────────
    for eid, items in all_news.items():
        entity = entities_map.get(eid)
        entity_name = entity.name if entity else eid
        entity_type = entity.entity_type.value if entity else "unknown"
        for idx, n in enumerate(items):
            fill = _fill(GREEN_LIGHT2) if idx % 2 == 0 else _fill(WHITE)
            _data_cell(ws, row, 1,  entity_name,             fill)
            _data_cell(ws, row, 2,  entity_type.capitalize(), fill)
            c = ws.cell(row=row, column=3, value="Accepted")
            c.font      = _font(bold=True, size=10, color=GREEN_MID2)
            c.fill      = fill
            c.border    = _border()
            c.alignment = _align()
            _data_cell(ws, row, 4,  "Passed all checks",     fill)
            _data_cell(ws, row, 5,  n.title,                 fill)
            _data_cell(ws, row, 6,  n.source,                fill)
            _data_cell(ws, row, 7,  "View",                  fill, hyperlink=n.url or n.original_url)
            _data_cell(ws, row, 8,  n.published_date,        fill)
            _data_cell(ws, row, 9,  n.period,                fill)
            _data_cell(ws, row, 10, n.primary_category,      fill)
            _data_cell(ws, row, 11, n.secondary_category or "", fill)
            _data_cell(ws, row, 12, n.topic_queried or "",   fill)
            _data_cell(ws, row, 13, n.fetch_source or "",    fill)
            ws.row_dimensions[row].height = 40
            row += 1

    # ── Rejected / duplicate entries from audit log ───────────────────────────
    non_accepted = [e for e in audit_entries if e.action != "accepted"]
    for idx, entry in enumerate(non_accepted):
        action = entry.action  # duplicate_removed | validation_rejected | url_invalid
        if action == "duplicate_removed":
            fill     = _fill(ORANGE_LIGHT) if idx % 2 == 0 else _fill(WHITE)
            status   = "Duplicate"
            s_color  = ORANGE_MID
        else:
            fill     = _fill(RED_LIGHT) if idx % 2 == 0 else _fill(WHITE)
            status   = action.replace("_", " ").title()
            s_color  = RED_MID

        # Prefer entity_type stored on the entry; fall back to live entities map
        entity      = entities_map.get(entry.entity_id)
        stored_type = (entry.entity_type or "").strip()
        if stored_type:
            entity_type = stored_type.capitalize()
        else:
            entity_type = entity.entity_type.value.capitalize() if entity else "Unknown"

        win_period = ""
        if getattr(entry, "window_from", "") and getattr(entry, "window_to", ""):
            win_period = f"{entry.window_from} → {entry.window_to}"

        _data_cell(ws, row, 1,  entry.entity_name,            fill)
        _data_cell(ws, row, 2,  entity_type,                   fill)
        c = ws.cell(row=row, column=3, value=status)
        c.font      = _font(bold=True, size=10, color=s_color)
        c.fill      = fill
        c.border    = _border()
        c.alignment = _align()
        _data_cell(ws, row, 4,  entry.reason,                  fill)
        _data_cell(ws, row, 5,  entry.article_title,           fill)
        _data_cell(ws, row, 6,  "",                            fill)  # publication source not in audit log
        _data_cell(ws, row, 7,  "View" if entry.source_url else "", fill,
                   hyperlink=entry.source_url or None)
        _data_cell(ws, row, 8,  entry.run_date,                fill)
        _data_cell(ws, row, 9,  win_period,                    fill)
        _data_cell(ws, row, 10, "",                            fill)
        _data_cell(ws, row, 11, "",                            fill)
        _data_cell(ws, row, 12, "",                            fill)
        _data_cell(ws, row, 13, entry.fetch_source or "",      fill)
        ws.row_dimensions[row].height = 40
        row += 1

    _copyright_row(ws, ncols, row + 1)
    ws.freeze_panes = "A4"


# ── Public entry point (standard digest) ──────────────────────────────────────
def generate_excel_report(
    client_data:   List[Dict],
    prospect_data: List[Dict],
    industry_data: List[Dict],
    window:        dict | None = None,
    gap_report:    dict | None = None,
    audit_entries: list | None = None,
    entities:      list | None = None,
) -> bytes:
    """
    Generates a multi-sheet .xlsx report and returns bytes for streaming.

    Sheets:
        1. Summary
        2. Client Report
        3. Prospect Report
        4. Industry News
        5. Gap Report (if gap_report provided)
        6. Pipeline Detail — all articles (accepted + rejected) with fetch API
           and run period (only added when audit_entries provided)
    """
    wb = openpyxl.Workbook()
    active = wb.active
    if active is not None:
        wb.remove(active)

    _build_summary(wb, client_data, prospect_data, industry_data, window)
    _build_client_sheet(wb,   client_data,   window)
    _build_prospect_sheet(wb, prospect_data, window)
    _build_industry_sheet(wb, industry_data, window)

    if gap_report:
        _build_gap_sheet(wb, gap_report)

    if audit_entries is not None:
        # Build entity map and all_news dict from the grouped data
        all_entities = [item["entity"] for item in client_data + prospect_data + industry_data]
        entities_map = {e.id: e for e in (entities or all_entities)}
        all_news = {}
        for item in client_data + prospect_data + industry_data:
            eid = item["entity"].id
            all_news[eid] = item["news"]
        _build_full_analytics(wb, entities_map, all_news, audit_entries, window)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _standardize_reason_xl(action: str, reason: str) -> str:
    """Inline version of the main.py helper — avoids circular import."""
    if action == "duplicate_removed":
        return "Duplicate Article"
    r = (reason or "").lower()
    if "before the window" in r or "too old" in r:
        return "Time Period Mismatch"
    if "future article" in r or "after the window end" in r:
        return "Date Validation Failure"
    if "parse" in r and "date" in r:
        return "Invalid Date Format"
    if "non-english" in r or "language field" in r:
        return "Non-English Content"
    if "domain blocked" in r or "social media" in r or "forum" in r:
        return "Source Validation Failure"
    if "not relevant" in r or "search drift" in r or "entity name not found" in r:
        return "Entity Name Mismatch"
    if "title is empty" in r or "[removed]" in r or "[deleted]" in r:
        return "Missing/Invalid Title"
    if action == "url_invalid" or ("url status" in r and "invalid" in r):
        return "Hyperlink Not Working"
    return "Validation Failure"


# ── Gap Report export ──────────────────────────────────────────────────────────
def generate_gap_report_excel(all_gap_reports: list, run_history: list | None = None) -> bytes:
    """
    Multi-sheet Excel report covering every historical gap report run.

    Sheet 1 — Run Summary: one row per run with aggregate stats
    Sheet 2 — Entity Summary: all entities across all runs
    Sheet 3 — No News Entities: entities that had zero articles per run
    Sheet 4 — Topic Gaps: per-entity missing topic detail
    """
    TEAL_DARK  = "0F5B5B"
    TEAL_MID   = "1A8585"
    TEAL_LIGHT = "E0F5F5"

    wb = openpyxl.Workbook()
    active = wb.active
    if active:
        wb.remove(active)

    # Sheet 1 — Run Summary
    ws1 = wb.create_sheet("Run Summary")
    hdrs1 = ["Run ID", "Run Date", "Started", "Period", "APIs Used", "Status",
             "Raw Fetched", "Dupes Removed", "Validation Rejected", "Accepted",
             "Acceptance %", "Validation Accuracy %", "Duration (s)"]
    ncols1 = len(hdrs1)
    _title_row(ws1, "Gap Report — Run Summary (All Runs)", ncols1, TEAL_DARK)
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols1)
    c = ws1.cell(row=2, column=1,
                 value=f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  {len(all_gap_reports)} run(s)")
    c.font = _font(size=10, color=WHITE); c.fill = _fill(TEAL_DARK); c.alignment = _center()
    _header_row(ws1, hdrs1, row=3, bg=TEAL_MID)
    _set_widths(ws1, [22, 14, 20, 28, 28, 14, 14, 16, 18, 14, 16, 20, 14])
    ws1.freeze_panes = "A4"

    sorted_reports = sorted(all_gap_reports, key=lambda r: r.get("started_at", r.get("run_date", "")), reverse=True)
    for idx, rpt in enumerate(sorted_reports):
        raw      = rpt.get("raw_fetched", 0)
        dupes    = rpt.get("duplicates_removed", 0)
        rejected = rpt.get("articles_rejected", 0)
        accepted = rpt.get("total_articles", 0)
        after_dd = max(0, raw - dupes)
        acc_pct  = round(accepted / raw * 100, 1) if raw else 0
        val_acc  = round(accepted / after_dd * 100, 1) if after_dd else 0
        fill = _fill(TEAL_LIGHT) if idx % 2 == 0 else _fill(WHITE)
        row = idx + 4
        period = f"{rpt.get('window', {}).get('from', '')} → {rpt.get('window', {}).get('to', '')}"
        apis   = ", ".join(rpt.get("api_sources_used", []))
        for ci, val in enumerate([
            rpt.get("run_id", ""), rpt.get("run_date", ""),
            rpt.get("started_at", ""), period, apis,
            rpt.get("status", "completed").title(),
            raw, dupes, rejected, accepted,
            acc_pct, val_acc, rpt.get("duration_seconds", ""),
        ], 1):
            _data_cell(ws1, row, ci, val, fill)
        ws1.row_dimensions[row].height = 18
    _copyright_row(ws1, ncols1, len(sorted_reports) + 5)

    # Sheet 2 — Entity Summary
    ws2 = wb.create_sheet("Entity Summary")
    hdrs2 = ["Run Date", "Run ID", "Entity Name", "Entity Type",
             "Raw Fetched", "Accepted", "Filtered", "Topics Covered", "Status"]
    _title_row(ws2, "Gap Report — Entity Summary (All Runs)", len(hdrs2), TEAL_DARK)
    _period_row(ws2, None, len(hdrs2), TEAL_MID)
    _header_row(ws2, hdrs2, row=3, bg=TEAL_MID)
    _set_widths(ws2, [14, 22, 28, 14, 14, 14, 14, 16, 16])
    ws2.freeze_panes = "A4"
    row2 = 4
    for rpt in sorted_reports:
        tg_map = rpt.get("topic_gaps", {})
        for eidx, item in enumerate(rpt.get("entity_summary", [])):
            name    = item.get("name", "")
            raw     = item.get("raw", 0)
            final   = item.get("final", 0)
            gaps    = tg_map.get(name, {}).get("missing_topics", [])
            covered = 12 - len(gaps)
            if final > 0:   status = "News Found"
            elif raw > 0:   status = "Filtered Out"
            else:           status = "No News"
            fill = _fill(TEAL_LIGHT) if eidx % 2 == 0 else _fill(WHITE)
            for ci, val in enumerate([
                rpt.get("run_date", ""), rpt.get("run_id", ""),
                name, item.get("type", ""),
                raw, final, max(0, raw - final),
                f"{covered}/12", status,
            ], 1):
                _data_cell(ws2, row2, ci, val, fill)
            ws2.row_dimensions[row2].height = 18
            row2 += 1
    _copyright_row(ws2, len(hdrs2), row2 + 1)

    # Sheet 3 — No News Entities
    ws3 = wb.create_sheet("No News Entities")
    hdrs3 = ["Run Date", "Run ID", "Period", "Entity Name", "Entity Type"]
    _title_row(ws3, "Gap Report — Entities with No News", len(hdrs3), AMBER_DARK)
    _period_row(ws3, None, len(hdrs3), AMBER_DARK)
    _header_row(ws3, hdrs3, row=3, bg=AMBER_DARK)
    _set_widths(ws3, [14, 22, 28, 28, 14])
    ws3.freeze_panes = "A4"
    row3 = 4
    for rpt in sorted_reports:
        period3 = f"{rpt.get('window', {}).get('from', '')} → {rpt.get('window', {}).get('to', '')}"
        for nidx, item in enumerate(rpt.get("no_news_at_all", [])):
            fill = _fill(AMBER) if nidx % 2 == 0 else _fill(WHITE)
            for ci, val in enumerate([
                rpt.get("run_date", ""), rpt.get("run_id", ""),
                period3, item.get("name", ""), item.get("type", ""),
            ], 1):
                _data_cell(ws3, row3, ci, val, fill)
            ws3.row_dimensions[row3].height = 18
            row3 += 1
    _copyright_row(ws3, len(hdrs3), row3 + 1)

    # Sheet 4 — Topic Gaps
    ws4 = wb.create_sheet("Topic Gaps")
    hdrs4 = ["Run Date", "Run ID", "Period", "Entity Name", "Entity Type", "Missing Topics", "Count Missing"]
    _title_row(ws4, "Gap Report — Topic-Level Gaps (All Runs)", len(hdrs4), AMBER_DARK)
    _period_row(ws4, None, len(hdrs4), AMBER_DARK)
    _header_row(ws4, hdrs4, row=3, bg=AMBER_DARK)
    _set_widths(ws4, [14, 22, 28, 28, 14, 60, 14])
    ws4.freeze_panes = "A4"
    row4 = 4
    for rpt in sorted_reports:
        period4 = f"{rpt.get('window', {}).get('from', '')} → {rpt.get('window', {}).get('to', '')}"
        for tidx, (ename, info) in enumerate(rpt.get("topic_gaps", {}).items()):
            fill    = _fill(AMBER) if tidx % 2 == 0 else _fill(WHITE)
            missing = info.get("missing_topics", [])
            for ci, val in enumerate([
                rpt.get("run_date", ""), rpt.get("run_id", ""),
                period4, ename, info.get("entity_type", ""),
                ", ".join(missing), len(missing),
            ], 1):
                _data_cell(ws4, row4, ci, val, fill)
            ws4.row_dimensions[row4].height = 30
            row4 += 1
    _copyright_row(ws4, len(hdrs4), row4 + 1)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ── Run History export ─────────────────────────────────────────────────────────
def generate_run_history_excel(runs: list) -> bytes:
    """Single-sheet run history with all stats and computed accuracy columns."""
    wb = openpyxl.Workbook()
    active = wb.active
    if active:
        wb.remove(active)

    ws = wb.create_sheet("Run History")
    hdrs = ["Run ID", "Started", "Completed", "Status", "Scope",
            "APIs Used", "Period", "Raw Fetched", "Dupes", "Rejected",
            "Accepted", "Acceptance %", "Validation Accuracy %",
            "AI Calls", "Tokens", "Duration (s)"]
    ncols = len(hdrs)
    _title_row(ws, "Run History — All Pipeline Runs", ncols, GRAY_DARK)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1,
                value=f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  {len(runs)} run(s)")
    c.font = _font(size=10, color=WHITE); c.fill = _fill(GRAY_DARK); c.alignment = _center()
    _header_row(ws, hdrs, row=3, bg=GRAY_DARK)
    _set_widths(ws, [22, 20, 20, 12, 20, 28, 28, 14, 10, 12, 12, 14, 20, 12, 14, 14])
    ws.freeze_panes = "A4"

    sorted_runs = sorted(runs, key=lambda r: r.get("started_at", ""), reverse=True)
    for idx, r in enumerate(sorted_runs):
        raw      = r.get("raw_fetched") or (r.get("total_articles", 0) + r.get("duplicates_removed", 0) + r.get("articles_rejected", 0))
        dupes    = r.get("duplicates_removed", 0)
        rejected = r.get("articles_rejected", 0)
        accepted = r.get("total_articles", 0)
        after_dd = max(0, raw - dupes)
        acc_pct  = round(accepted / raw * 100, 1) if raw else 0
        val_acc  = round(accepted / after_dd * 100, 1) if after_dd else 0
        tokens   = (r.get("prompt_tokens", 0) or 0) + (r.get("completion_tokens", 0) or 0)
        period   = f"{r.get('window_from', '')} → {r.get('window_to', '')}" if r.get("window_from") else r.get("window_label", "")
        apis     = ", ".join(r.get("api_sources_used", []))
        scope    = r.get("entity_name", f"All ({r.get('total_entities', '')})")
        fill     = _fill(GRAY_LIGHT) if idx % 2 == 0 else _fill(WHITE)
        row = idx + 4
        for ci, val in enumerate([
            r.get("run_id", ""), r.get("started_at", ""), r.get("completed_at", ""),
            r.get("status", "").title(), scope, apis, period,
            raw, dupes, rejected, accepted, acc_pct, val_acc,
            r.get("ai_calls", ""), tokens, r.get("duration_seconds", ""),
        ], 1):
            _data_cell(ws, row, ci, val, fill)
        ws.row_dimensions[row].height = 18
    _copyright_row(ws, ncols, len(sorted_runs) + 5)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ── Audit Report export ────────────────────────────────────────────────────────
def generate_audit_report_excel(audit_entries: list) -> bytes:
    """
    Three-sheet audit report:
    Sheet 1 — All entries
    Sheet 2 — Accepted only
    Sheet 3 — Rejected / Duplicates with standardized reason
    """
    RED_LIGHT = "FEE2E2"; RED_MID = "B91C1C"
    GRN_LIGHT = "DCFCE7"; GRN_MID = "166534"
    ORG_LIGHT = "FEF9C3"; ORG_MID = "B45309"

    wb = openpyxl.Workbook()
    active = wb.active
    if active:
        wb.remove(active)

    common_hdrs = ["Run Date", "Period", "Entity Name", "Entity Type",
                   "Action", "Std. Reason", "API Source", "Article Title", "Reason Detail", "Source URL"]
    cwidths = [12, 26, 24, 14, 16, 24, 12, 44, 44, 16]

    def _write_sheet(ws, title: str, entries, bg_dark, bg_mid, row_fill_even, row_fill_odd):
        ncols = len(common_hdrs)
        _title_row(ws, title, ncols, bg_dark)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c2 = ws.cell(row=2, column=1,
                     value=f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  {len(entries)} entries")
        c2.font = _font(size=10, color=WHITE); c2.fill = _fill(bg_dark); c2.alignment = _center()
        _header_row(ws, common_hdrs, row=3, bg=bg_mid)
        _set_widths(ws, cwidths)
        ws.freeze_panes = "A4"
        for idx, e in enumerate(entries):
            period = (f"{e.window_from} → {e.window_to}" if getattr(e, "window_from", "") and getattr(e, "window_to", "") else "")
            std    = _standardize_reason_xl(e.action, e.reason)
            fill   = _fill(row_fill_even) if idx % 2 == 0 else _fill(row_fill_odd)
            row    = idx + 4
            for ci, val in enumerate([
                e.run_date, period, e.entity_name,
                (e.entity_type or "").title(), e.action.replace("_", " ").title(),
                std, e.fetch_source or "", e.article_title,
                e.reason, e.source_url or "",
            ], 1):
                _data_cell(ws, row, ci, val, fill)
            ws.row_dimensions[row].height = 32
        _copyright_row(ws, len(common_hdrs), len(entries) + 5)

    # Try to handle both AuditEntry objects and dicts
    def _as_obj(e):
        if hasattr(e, "action"):
            return e
        from types import SimpleNamespace
        return SimpleNamespace(**e)

    entries_obj  = [_as_obj(e) for e in audit_entries]
    accepted_obj = [e for e in entries_obj if e.action == "accepted"]
    rejected_obj = [e for e in entries_obj if e.action != "accepted"]

    _write_sheet(wb.create_sheet("All Entries"),  "Audit Log — All Entries",  entries_obj,  GRAY_DARK, "5F5E5A", GRAY_LIGHT, WHITE)
    _write_sheet(wb.create_sheet("Accepted"),     "Audit Log — Accepted",     accepted_obj, GREEN_DARK, GREEN_MID, GREEN_LIGHT, WHITE)
    _write_sheet(wb.create_sheet("Rejected"),     "Audit Log — Rejected & Duplicates", rejected_obj, RED_MID, RED_MID, RED_LIGHT, WHITE)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ── Validation Report export ───────────────────────────────────────────────────
def generate_validation_report_excel(entities: list, all_news: dict, audit_entries: list) -> bytes:
    """
    Three-sheet validation report:
    Sheet 1 — Summary by rejection reason
    Sheet 2 — Accepted articles with URL status
    Sheet 3 — Rejected articles with standardized reason
    """
    RED_LIGHT = "FEE2E2"; RED_MID = "B91C1C"
    GRN_LIGHT = "DCFCE7"; GRN_MID = "166534"
    BLUE_MID2 = "2563EB"

    wb = openpyxl.Workbook()
    active = wb.active
    if active:
        wb.remove(active)

    entities_map = {e.id: e for e in entities}

    # ── Sheet 1: Reason Summary ───────────────────────────────────────────────
    ws1 = wb.create_sheet("Reason Summary")
    _title_row(ws1, "Validation Report — Rejection Reason Summary", 4, GRAY_DARK)
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    c2 = ws1.cell(row=2, column=1,
                  value="How validation accuracy is calculated: Accepted ÷ (Raw − Duplicates) × 100. "
                        "Duplicates are URL/headline matches removed before validation. "
                        "Soft-tagged articles (limited preview, paywall) pass through and are counted as Accepted.")
    c2.font = _font(size=9, color=WHITE); c2.fill = _fill(GRAY_DARK); c2.alignment = _align(); c2.alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[2].height = 36
    _header_row(ws1, ["Standardized Reason", "Count", "% of Non-Accepted", "Description"], row=3, bg=GRAY_DARK)
    _set_widths(ws1, [32, 12, 20, 60])
    ws1.freeze_panes = "A4"

    def _as_obj(e):
        if hasattr(e, "action"):
            return e
        from types import SimpleNamespace
        return SimpleNamespace(**e)

    entries_obj = [_as_obj(e) for e in audit_entries]
    non_accepted = [e for e in entries_obj if e.action != "accepted"]
    reason_counts: dict = {}
    for e in non_accepted:
        label = _standardize_reason_xl(e.action, e.reason)
        reason_counts[label] = reason_counts.get(label, 0) + 1

    _REASON_DESC = {
        "Duplicate Article":         "URL or headline matched an existing article — removed before validation.",
        "Time Period Mismatch":       "Article published before the selected date window start.",
        "Date Validation Failure":    "Article published after the window end (future-dated) or date couldn't be parsed.",
        "Invalid Date Format":        "Published date string could not be interpreted.",
        "Non-English Content":        "Article language field explicitly set to a non-English language.",
        "Source Validation Failure":  "Article URL is from a blocked domain (social media, forum).",
        "Entity Name Mismatch":       "Entity name not found (capitalised) in title or content — search drift.",
        "Missing/Invalid Title":      "Title is empty, null, or marked as [Removed] / [Deleted].",
        "Hyperlink Not Working":      "URL returned HTTP 404 / 410 or DNS failure after validation.",
        "Validation Failure":         "Other validation rule triggered.",
    }
    total_non = len(non_accepted) or 1
    for ridx, (label, cnt) in enumerate(sorted(reason_counts.items(), key=lambda x: -x[1]), 4):
        fill = _fill(GRAY_LIGHT) if ridx % 2 == 0 else _fill(WHITE)
        for ci, val in enumerate([label, cnt, f"{round(cnt/total_non*100,1)}%", _REASON_DESC.get(label, "")], 1):
            _data_cell(ws1, ridx, ci, val, fill)
        ws1.row_dimensions[ridx].height = 22
    _copyright_row(ws1, 4, len(reason_counts) + 6)

    # ── Sheet 2: Accepted Articles ────────────────────────────────────────────
    ws2 = wb.create_sheet("Accepted Articles")
    hdrs2 = ["Entity Name", "Entity Type", "Article Title", "Source", "Published",
             "Period", "URL Status", "Fetch API", "Primary Category", "URL"]
    _title_row(ws2, "Validation Report — Accepted Articles", len(hdrs2), GREEN_DARK)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(hdrs2))
    c2b = ws2.cell(row=2, column=1,
                   value=f"Accepted articles include url_status=ok, redirect, unknown, and paywall-tagged articles.")
    c2b.font = _font(size=10, color=WHITE); c2b.fill = _fill(GREEN_MID); c2b.alignment = _center()
    _header_row(ws2, hdrs2, row=3, bg=GREEN_MID)
    _set_widths(ws2, [24, 14, 44, 20, 14, 20, 14, 14, 30, 14])
    ws2.freeze_panes = "A4"
    row2 = 4
    for eid, items in all_news.items():
        entity = entities_map.get(eid)
        ename  = entity.name if entity else eid
        etype  = entity.entity_type.value if entity else "unknown"
        for nidx, n in enumerate(items):
            fill = _fill(GRN_LIGHT) if nidx % 2 == 0 else _fill(WHITE)
            for ci, val in enumerate([
                ename, etype.title(), n.title, n.source, n.published_date,
                n.period, n.url_status or "ok", n.fetch_source or "",
                n.primary_category or "", "",
            ], 1):
                if ci == 10:
                    _data_cell(ws2, row2, ci, "View", fill, hyperlink=n.url or n.original_url)
                else:
                    _data_cell(ws2, row2, ci, val, fill)
            ws2.row_dimensions[row2].height = 30
            row2 += 1
    _copyright_row(ws2, len(hdrs2), row2 + 1)

    # ── Sheet 3: Rejected Articles ────────────────────────────────────────────
    ws3 = wb.create_sheet("Rejected Articles")
    hdrs3 = ["Run Date", "Entity Name", "Entity Type", "Action",
             "Standardized Reason", "API Source", "Article Title", "Reason Detail", "URL"]
    _title_row(ws3, "Validation Report — Rejected Articles", len(hdrs3), RED_MID)
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(hdrs3))
    c3b = ws3.cell(row=2, column=1,
                   value=f"Rejected articles: duplicates + validation failures + invalid hyperlinks.")
    c3b.font = _font(size=10, color=WHITE); c3b.fill = _fill(RED_MID); c3b.alignment = _center()
    _header_row(ws3, hdrs3, row=3, bg=RED_MID)
    _set_widths(ws3, [12, 24, 14, 18, 26, 12, 44, 44, 14])
    ws3.freeze_panes = "A4"
    row3 = 4
    for nidx, e in enumerate(non_accepted):
        fill = _fill(RED_LIGHT) if nidx % 2 == 0 else _fill(WHITE)
        std  = _standardize_reason_xl(e.action, e.reason)
        for ci, val in enumerate([
            e.run_date, e.entity_name, (e.entity_type or "").title(),
            e.action.replace("_", " ").title(), std,
            e.fetch_source or "", e.article_title, e.reason, "",
        ], 1):
            if ci == 9:
                _data_cell(ws3, row3, ci, "View" if e.source_url else "", fill,
                           hyperlink=e.source_url if e.source_url else None)
            else:
                _data_cell(ws3, row3, ci, val, fill)
        ws3.row_dimensions[row3].height = 30
        row3 += 1
    _copyright_row(ws3, len(hdrs3), row3 + 1)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def generate_full_analytics_report(
    entities:      list,
    all_news:      dict,
    audit_entries: list,
    window:        dict | None = None,
) -> bytes:
    """
    Generates a single-sheet .xlsx with every article (accepted + rejected).
    entities: List[Entity]
    all_news: Dict[entity_id, List[NewsItem]]
    audit_entries: List[AuditEntry]
    window: date range dict with 'label', 'from', 'to' keys
    """
    entities_map = {e.id: e for e in entities}

    wb = openpyxl.Workbook()
    active = wb.active
    if active is not None:
        wb.remove(active)

    _build_full_analytics(wb, entities_map, all_news, audit_entries, window)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()